import pandas as pd
import numpy as np
import xlrd
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
import datetime
#The begining of this program will create a dataframes per company with just the necessary data that is needed.


def createDFSource(locSource):#The Sheet name is given at all times if you need a different sheet this is where you change it. You will also need to change findDate function
    sourceDF = pd.read_excel(locSource, sheet_name = 'TT2120190024', header = 8)
    sourceDF2 = sourceDF[['SLIN', 'TYPE', 'CATEGORY', 'Current Hours', 'Current Amount']] #These are the only columns we will need
    return sourceDF2

def createDFFinal(sourceDF): #Here we pare the DF down to what we need to append to the Hours sheet in FY 2019
    sourceDF = sourceDF[['CATEGORY', 'Current Hours']].dropna() #drop all not a number values in Category and Current Hours
    dropArray = []
    for index, row in sourceDF.iterrows():
        if '0' in row["CATEGORY"] or 'LLC' in row["CATEGORY"] or "Inc." in row["CATEGORY"] or "Technologies" in row["CATEGORY"]: #If there is a company we don't care about it. Only the Category
            dropArray.append(index) #add the index to a list so that we can drop it later
        if row['Current Hours'] == 0: #anything that has zero hours we don't need to add to the data frame
            dropArray.append(index)

    sourceDF.drop(dropArray, inplace = True)
    return sourceDF

def createDFHours(locDest): #read in the data frame that we will edit from the FY 2019 Hours sheet
    destDFHours = pd.read_excel(locDest , sheet_name = 'Hours', header = 3)
    dfIndex = 0

    for row in destDFHours["Labor Category"]: #at the first row of nan we want to cut it
        if row is np.nan:
            return destDFHours[:dfIndex]
        else:
            dfIndex += 1
    return -1

def updateDFHours(sourceDF , finalDFHours, locSource): #This returns the changed dataFrame that used to be FY 2019
    month = findDate(locSource)
    month = month.split('-')[1].split('/')[0]

    for index, row in sourceDF.iterrows(): #iterate through the rows in the Source File

        hourRow = finalDFHours.loc[finalDFHours["Labor Category"] == row["CATEGORY"]].copy() #save the row where the Labor Category (Final Excel Sheet) = the Category(Source Excel Sheet)

        hourRow["Worked." + month] = row["Current Hours"]

        finalDFHours.loc[finalDFHours["Labor Category"] == row["CATEGORY"]] = hourRow
    return finalDFHours

def findDate(locSource): #hard coded where the date is in this spreadsheet and the sheet that is needed
    workbook = xlrd.open_workbook(locSource)
    sheet = workbook.sheet_by_name('TT2120190024')
    date = sheet.cell_value(rowx=6, colx=1)
    return date

#Code written by MaxU on stack overflow. Thanks for the help :)
# ignore [engine] parameter if it was passed
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs, index = False)

    # save the workbook
    writer.save()
#End of code written by MaxU

#The first sheet is now completed. Now we move on to the second sheet

def travelTotal(sourceDF2): #rounding error will be taken care of at the location this is being used
    travelAmount = sum(list(sourceDF2.loc[sourceDF2["TYPE" ] == "Travel"]['Current Amount']))
    return travelAmount

def calcTotal(sourceDF2): #used decimal due to this being a financial document
    travel = Decimal(travelTotal(sourceDF2))
    currAmount = Decimal(sourceDF2["Current Amount"].iloc[-1])
    total = Decimal(currAmount-travel)
    roundedTotal = Decimal(total.quantize(Decimal('.01'), rounding=ROUND_HALF_UP))
    return roundedTotal

def reimbTotal (reimbSource): #the Cust Ord is hard coded because it's his personal employee number.
    reimbDF = pd.read_excel(reimbSource, header = 0)
    reimbDF = reimbDF[['TOTAL EXP IN MON', 'CUST ORD']]
    final = 0

    for index, row in reimbDF.iterrows():
        if row['CUST ORD'] == 'HQ0147966843':
            final = row['TOTAL EXP IN MON']

    return final

def createDFCOL(locDest):
    destDFCOL = pd.read_excel(locDest, sheet_name = 'COL Report', header = 1)
    dfIndex = 0
    for row in destDFCOL["Month"]: #at the first row of nan we want to cut it
        if row is np.nan:
            return destDFCOL[:dfIndex-2] #I subtracted two here because I don't need Travel or Total row's in the dataframe
        else:
            dfIndex += 1
    return -1

def neededDFCOL(destDFCOL):
    destDFCOL = destDFCOL[['Month', 'Monthly Accruals.1', 'Monthly Accruals.2', 'Monthly Accruals.3']] #I only need to change data in these columns
    return destDFCOL

def formatDate(locSource):
    date = findDate(locSource)
    date = date.split('-')[1]

    date = datetime.datetime.strptime(date, "%m/%d/%y")

    return date

def findIndex(dfCOL, locSource):
    dateFind = formatDate(locSource)
    index = 0
    for index, row in dfCOL.iterrows():
        dateActual = row['Month']
        if dateFind.month == dateActual.month and dateFind.year == dateActual.year:
            return index

def changeColReport(dfCOL, sourceDF, locSource, locDest, locReimb):
    travel = travelTotal(sourceDF)
    total = calcTotal(sourceDF)
    reimb = reimbTotal(locReimb)
    index = findIndex(dfCOL, locSource) + 3 #add 3 because the data fram begins at 3 lower than the excel sheet does

    wb = load_workbook(locDest)
    ws = wb['COL Report']
    ws['c' + str(index)] = total
    ws['J' + str(index)] = reimb
    ws['Q' + str(index)] = travel
    wb.save(locDest)

if __name__ == "__main__":
    locSource = input("Enter the name of the Torch report (Make sure you have the .xlsx on the end of it)")
    locDest = input("Enter the name of the FY excel sheet")
    locReimb = input("Enter the name of the Reimbursement sheet")

    sourceDF = createDFSource(locSource)
    sourceDFFinal = createDFFinal(sourceDF)
    destDFHours = createDFHours(locDest)
    destDFHours = updateDFHours(sourceDFFinal, destDFHours, locSource)

    append_df_to_excel(locDest, destDFHours, sheet_name = 'Hours', startrow = 3 )

    colDF = createDFCOL(locDest)
    colDF = neededDFCOL(colDF)
    changeColReport(colDF, sourceDF, locSource, locDest, locReimb)

    print("All done.")
